import os
from datetime import date
from tabulate import tabulate as tab

import pandas as pd

import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
from matplotlib.collections import PatchCollection
from mpl_toolkits.basemap import Basemap

import numpy as np

from PIL import Image, ImageOps

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

today = date.today()

shp_simple_countries = r'C:/Users/Georges/PycharmProjects/data/simple_countries/simple_countries'
shp_simple_areas = r'C:/Users/Georges/PycharmProjects/data/simple_areas/simple_areas'
inputCountryConversion = r'C:/Users/Georges/PycharmProjects/data/countries_conversion.xlsx'

workDirectory = r'C:/Users/Georges/Downloads/Webinar/'

WebinarFileName = '20210323_Webinar_Russian'

ReportExcelFile = workDirectory + WebinarFileName + '_Report.xlsx'
NewAddThenDeleteExcelFile = workDirectory + WebinarFileName + '_NewAddJooThenDelete.xlsx'
NewCollectExcelFile = workDirectory + WebinarFileName + '_NewToCollect.xlsx'


# WEBINAR EXCEL IMPORT
inputExcelFile = workDirectory+WebinarFileName+'.xlsx'
df = pd.read_excel(inputExcelFile, sheet_name='export', engine='openpyxl',
                   usecols=['ID', 'Honorary title', 'First name', 'Last name', 'Email', 'Live Location:Country',
                            'Industries:Industries', 'Business/Professional sector', 'How did you hear about AMS?'
                            ])

# REMOVE ADMIN AND TEAM
index_drop1 = df[df['Email'].apply(lambda x: x.endswith('@informa.com'))].index
df = df.drop(index_drop1)
index_drop2 = df[df['Email'].apply(lambda x: x.endswith('@euromedicom.com'))].index
df = df.drop(index_drop2)
index_drop3 = df[df['Email'].apply(lambda x: x.endswith('@eurogin.com'))].index
df = df.drop(index_drop3)
index_drop4 = df[df['Email'].apply(lambda x: x.endswith('@multispecialtysociety.com'))].index
df = df.drop(index_drop4)
index_drop5 = df[df['Email'].apply(lambda x: x.endswith('@ce.com.co'))].index
df = df.drop(index_drop5)
index_drop6 = df[df['Email'].apply(lambda x: x == ('max.carter11@yahoo.com'))].index
df = df.drop(index_drop6)
index_drop7 = df[df['Email'].apply(lambda x: x == ('eo-wilson@biodiv.us'))].index
df = df.drop(index_drop7)
index_drop8 = df[df['Email'].apply(lambda x: x == ('cartoperso@yahoo.fr'))].index
df = df.drop(index_drop8)
index_drop9 = df[df['Email'].apply(lambda x: x == ('georges.hinot@gmail.com'))].index
df = df.drop(index_drop9)
index_drop10 = df[df['Email'].apply(lambda x: x == ('hakimislim@yahoo.fr'))].index
df = df.drop(index_drop10)
index_drop11 = df[df['Email'].apply(lambda x: x == ('salimsanslea@ymail.com'))].index
df = df.drop(index_drop11)
index_drop12 = df[df['Email'].apply(lambda x: x == ('salimsanslea@hotmail.com'))].index
df = df.drop(index_drop12)

# COUNT ALL PARTICIPANTS
participants = df.shape[0]


# JOO_ACYMAILING_SUBSCRIBER IMPORT
df_subscriber = pd.read_csv(workDirectory+'joo_acymailing_subscriber.csv', sep=',', usecols=['source', 'email'])
# SOURCES HARMONIZATION
df_subscriber['source'] = df_subscriber['source'].replace({'EXTERN: ': ''}, regex=True)
df_subscriber['source'] = df_subscriber['source'].replace({'PROSPECT: ': ''}, regex=True)


# COUNTRY CONVERSION IMPORT
df_CountryConversion = pd.read_excel(inputCountryConversion, sheet_name='countries', engine='openpyxl',
                   usecols=['COUNTRY_HB', 'continent_stat'])


# COUNT COUNTRY
df_Country_count = pd.DataFrame(df.groupby(['Live Location:Country'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Country_count = df_Country_count.fillna('Unknow')

df_Country_count['Percent'] = (df_Country_count['Total'] / df_Country_count['Total'].sum()) * 100
df_Country_count['Percent'] = df_Country_count['Percent'].round(decimals=1)


# COUNT AREAS
# JOIN LEFT WITH COUNTRY CONVERSION
df_WebinarAreas = pd.merge(df, df_CountryConversion, left_on='Live Location:Country', right_on='COUNTRY_HB', how='left')\
    [['Email', 'continent_stat']]

df_AreasCount = pd.DataFrame(df_WebinarAreas.groupby(['continent_stat'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_AreasCount = df_AreasCount.fillna('Unknow')

df_AreasCount['Percent'] = (df_AreasCount['Total'] / df_AreasCount['Total'].sum()) * 100
df_AreasCount['Percent'] = df_AreasCount['Percent'].round(decimals=1)


# COUNT CATEGORIES (CUSTOM FIELD Business/Professional sector)
df['Categories'] = df['Business/Professional sector'].str.split(': ').str[0]
df_Categories_count = pd.DataFrame(df.groupby(['Categories'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Categories_count = df_Categories_count.fillna('Unknow')

df_Categories_count['Percent'] = (df_Categories_count['Total'] / df_Categories_count['Total'].sum()) * 100
df_Categories_count['Percent'] = df_Categories_count['Percent'].round(decimals=1)


# COUNT SPECIALTIES (CUSTOM FIELD Business/Professional sector)
df['Specialties'] = df['Business/Professional sector'].str.split(': ').str[1]
df_Specialties_count = pd.DataFrame(df.groupby(['Specialties'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Specialties_count = df_Specialties_count.fillna('Unknow')

df_Specialties_count['Percent'] = (df_Specialties_count['Total'] / df_Specialties_count['Total'].sum()) * 100
df_Specialties_count['Percent'] = df_Specialties_count['Percent'].round(decimals=1)


# COUNT SPECIALTIES PER COUNTRY
df_SpecialtiesPerCountry_count = pd.DataFrame(df.groupby(['Live Location:Country', 'Specialties'], dropna=False)\
    .size(), columns=['Total']).sort_values(['Live Location:Country', 'Total'], ascending=[True, False]).reset_index()
df_SpecialtiesPerCountry_count = df_SpecialtiesPerCountry_count.fillna('Unknow')

df_SpecialtiesPerCountry_count['Percent'] = (df_SpecialtiesPerCountry_count['Total'] / df_SpecialtiesPerCountry_count['Total'].sum()) * 100
df_SpecialtiesPerCountry_count['Percent'] = df_SpecialtiesPerCountry_count['Percent'].round(decimals=2)


# COUNT EXPERTISE & INTERESTS (CUSTOM FIELD Industries:Industries)
df_tempIndustries = pd.DataFrame(pd.melt(df['Industries:Industries'].str.split(',', expand=True))['value'])
df_Industries_count = pd.DataFrame(df_tempIndustries.groupby(['value'], dropna=False).size(), columns=['Total'])\
    .reset_index()
df_Industries_count = df_Industries_count.fillna('AZERTY')

df_Industries_count['Percent'] = (df_Industries_count['Total'] / df.shape[0]) * 100
df_Industries_count['Percent'] = df_Industries_count['Percent'].round(decimals=2)

# EMPTY VALUES
industriesEmpty = df['Industries:Industries'].isna().sum()
industriesEmptyPercent = round((industriesEmpty / df.shape[0]) * 100, 2)

# REPLACE EMPTY VALUES AND SORT
df_Industries_count.loc[(df_Industries_count['value'] == 'AZERTY')] = [['Unknow', industriesEmpty, industriesEmptyPercent]]
df_Industries_count = df_Industries_count.sort_values(['Total'], ascending=False)


# COUNT HOW DID YOU HEAR ABOUT AMS (CUSTOM FIELD How did you hear about AMS?)
df_HowDidYouHearAboutUs_count = pd.DataFrame(df.groupby(['How did you hear about AMS?'], dropna=True).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_HowDidYouHearAboutUs_count = df_HowDidYouHearAboutUs_count.fillna('Unknow')

df_HowDidYouHearAboutUs_count['Percent'] = (df_HowDidYouHearAboutUs_count['Total'] / df_HowDidYouHearAboutUs_count['Total'].sum()) * 100
df_HowDidYouHearAboutUs_count['Percent'] = df_HowDidYouHearAboutUs_count['Percent'].round(decimals=1)

# REPLACE SOME VALUES
df_HowDidYouHearAboutUs_count['How did you hear about AMS?'] = df_HowDidYouHearAboutUs_count['How did you hear about AMS?'].replace(['Other: please specify'],'Other')


# COUNT SOURCES
# JOIN LEFT WITH SUBSCRIBERS
df_WebinarSubscriber = pd.merge(df, df_subscriber, left_on='Email', right_on='email', how='left')\
    [['Email', 'source']]

df_Sources = pd.DataFrame(df_WebinarSubscriber.groupby(['source'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Sources = df_Sources.fillna('')

df_Sources['Percent'] = (df_Sources['Total'] / df_Sources['Total'].sum()) * 100
df_Sources['Percent'] = df_Sources['Percent'].round(decimals=1)

# NEW EMAILS
df_WebinarNew = pd.DataFrame(df[~df['Email'].isin(df_subscriber['email'])])
newWebinar = df_WebinarNew.shape[0]


# COUNT NEW COUNTRY
df_NewCountry_count = pd.DataFrame(df_WebinarNew.groupby(['Live Location:Country'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_NewCountry_count = df_NewCountry_count.fillna('Unknow')

df_NewCountry_count['Percent'] = (df_NewCountry_count['Total'] / df_NewCountry_count['Total'].sum()) * 100
df_NewCountry_count['Percent'] = df_NewCountry_count['Percent'].round(decimals=1)


# COUNT NEW AREAS
# JOIN LEFT WITH COUNTRY CONVERSION
df_WebinarNewAreas = pd.merge(df_WebinarNew, df_CountryConversion, left_on='Live Location:Country', right_on='COUNTRY_HB', how='left')\
    [['Email', 'continent_stat']]

df_NewAreasCount = pd.DataFrame(df_WebinarNewAreas.groupby(['continent_stat'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_NewAreasCount = df_NewAreasCount.fillna('Unknow')

df_NewAreasCount['Percent'] = (df_NewAreasCount['Total'] / df_NewAreasCount['Total'].sum()) * 100
df_NewAreasCount['Percent'] = df_NewAreasCount['Percent'].round(decimals=1)


# COUNT NEW CATEGORIES (CUSTOM FIELD Business/Professional sector)
df_WebinarNew['Categories'] = df_WebinarNew['Business/Professional sector'].str.split(': ').str[0]
df_NewCategories_count = pd.DataFrame(df_WebinarNew.groupby(['Categories'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_NewCategories_count = df_NewCategories_count.fillna('Unknow')

df_NewCategories_count['Percent'] = (df_NewCategories_count['Total'] / df_NewCategories_count['Total'].sum()) * 100
df_NewCategories_count['Percent'] = df_NewCategories_count['Percent'].round(decimals=1)


# COUNT NEW SPECIALTIES (CUSTOM FIELD Business/Professional sector)
df_WebinarNew['Specialties'] = df_WebinarNew['Business/Professional sector'].str.split(': ').str[1]
df_NewSpecialties_count = pd.DataFrame(df_WebinarNew.groupby(['Specialties'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_NewSpecialties_count = df_NewSpecialties_count.fillna('Unknow')

df_NewSpecialties_count['Percent'] = (df_NewSpecialties_count['Total'] / df_NewSpecialties_count['Total'].sum()) * 100
df_NewSpecialties_count['Percent'] = df_NewSpecialties_count['Percent'].round(decimals=1)


# COUNT NEW SPECIALTIES PER COUNTRY
df_NewSpecialtiesPerCountry_count = pd.DataFrame(df_WebinarNew.groupby(['Live Location:Country', 'Specialties'], dropna=False)\
    .size(), columns=['Total']).sort_values(['Live Location:Country', 'Total'], ascending=[True, False]).reset_index()
df_NewSpecialtiesPerCountry_count = df_NewSpecialtiesPerCountry_count.fillna('Unknow')

df_NewSpecialtiesPerCountry_count['Percent'] = (df_NewSpecialtiesPerCountry_count['Total'] / df_NewSpecialtiesPerCountry_count['Total'].sum()) * 100
df_NewSpecialtiesPerCountry_count['Percent'] = df_NewSpecialtiesPerCountry_count['Percent'].round(decimals=2)


# COUNT NEW EXPERTISE & INTERESTS (CUSTOM FIELD Industries:Industries)
df_NewtempIndustries = pd.DataFrame(pd.melt(df_WebinarNew['Industries:Industries'].str.split(',', expand=True))['value'])
df_NewIndustries_count = pd.DataFrame(df_NewtempIndustries.groupby(['value'], dropna=False).size(), columns=['Total'])\
    .reset_index()
df_NewIndustries_count = df_NewIndustries_count.fillna('AZERTY')

df_NewIndustries_count['Percent'] = (df_NewIndustries_count['Total'] / df_WebinarNew.shape[0]) * 100
df_NewIndustries_count['Percent'] = df_NewIndustries_count['Percent'].round(decimals=2)

# EMPTY VALUES
industriesEmpty = df_WebinarNew['Industries:Industries'].isna().sum()
industriesEmptyPercent = round((industriesEmpty / df_WebinarNew.shape[0]) * 100, 2)

# REPLACE EMPTY VALUES AND SORT
df_NewIndustries_count.loc[(df_NewIndustries_count['value'] == 'AZERTY')] = [['Unknow', industriesEmpty, industriesEmptyPercent]]
df_NewIndustries_count = df_NewIndustries_count.sort_values(['Total'], ascending=False)


# COUNT NEW HOW DID YOU HEAR ABOUT AMS (CUSTOM FIELD How did you hear about AMS?)
df_NewHowDidYouHearAboutUs_count = pd.DataFrame(df_WebinarNew.groupby(['How did you hear about AMS?'], dropna=True).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_NewHowDidYouHearAboutUs_count = df_NewHowDidYouHearAboutUs_count.fillna('Unknow')

df_NewHowDidYouHearAboutUs_count['Percent'] = (df_NewHowDidYouHearAboutUs_count['Total'] / df_NewHowDidYouHearAboutUs_count['Total'].sum()) * 100
df_NewHowDidYouHearAboutUs_count['Percent'] = df_NewHowDidYouHearAboutUs_count['Percent'].round(decimals=1)

# REPLACE SOME VALUES
df_NewHowDidYouHearAboutUs_count['How did you hear about AMS?'] = df_NewHowDidYouHearAboutUs_count['How did you hear about AMS?'].replace(['Other: please specify'],'Other')


# EXCEL FILE: NEW TO ADD IN JOO_ACYMAILING_SUBSCRIBER THEN DELETE
df_WebinarNew['source'] = 'AMS '+str(today.strftime('%b %Y')).upper()
writer = pd.ExcelWriter(NewAddThenDeleteExcelFile, engine='xlsxwriter')
df_WebinarNew[['Email', 'source']].to_excel(writer, index=False, sheet_name='New Add Then Delete')
writer.save()


# EXCEL FILE: NEW TO COLLECT
df_WebinarNew = df_WebinarNew.drop(columns=['ID', 'source', 'Categories', 'Specialties'])
writer = pd.ExcelWriter(NewCollectExcelFile, engine='xlsxwriter')
df_WebinarNew.to_excel(writer, index=False, sheet_name='New Collect')
writer.save()


# EXCEL FILE: REPORT
writer = pd.ExcelWriter(ReportExcelFile, engine='xlsxwriter')

df_Country_count.to_excel(writer, index=False, sheet_name='Countries', header=['Country', 'Total', '%'])
df_AreasCount.to_excel(writer, index=False, sheet_name='Areas', header=['Area', 'Total', '%'])
df_Categories_count.to_excel(writer, index=False, sheet_name='Categories', header=['Category', 'Total', '%'])
df_Specialties_count.to_excel(writer, index=False, sheet_name='Specialties', header=['Specialty', 'Total', '%'])
df_SpecialtiesPerCountry_count.to_excel(writer, index=False, sheet_name='Specialties per country', header=['Country', 'Specialty', 'Total', '%'])
df_Industries_count.to_excel(writer, index=False, sheet_name='Expertise & Interests', header=['Expertise or Interest', 'Total', '%'])
df_HowDidYouHearAboutUs_count.to_excel(writer, index=False, sheet_name='How Did You Hear', header=['How did you hear about AMS (known)', 'Total', '%'])

df_Sources.to_excel(writer, index=False, sheet_name='Sources', header=['Source', 'Total', '%'])

df_NewCountry_count.to_excel(writer, index=False, sheet_name='News Countries', header=['Country', 'Total', '%'])
df_NewAreasCount.to_excel(writer, index=False, sheet_name='News Areas', header=['Area', 'Total', '%'])
df_NewCategories_count.to_excel(writer, index=False, sheet_name='News Categories', header=['Category', 'Total', '%'])
df_NewSpecialties_count.to_excel(writer, index=False, sheet_name='News Specialties', header=['Specialty', 'Total', '%'])
df_NewSpecialtiesPerCountry_count.to_excel(writer, index=False, sheet_name='News Specialties per country', header=['Country', 'Specialty', 'Total', '%'])
df_NewIndustries_count.to_excel(writer, index=False, sheet_name='News Expertise & Interests', header=['Expertise or Interest', 'Total', '%'])
df_NewHowDidYouHearAboutUs_count.to_excel(writer, index=False, sheet_name='News How Did You Hear', header=['How did you hear about AMS (known)', 'Total', '%'])

writer.save()

# EXCEL FILTERS
workbook = openpyxl.load_workbook(ReportExcelFile)
sheetsLits = workbook.sheetnames

for sheet in sheetsLits:
    worksheet = workbook[sheet]
    FullRange = 'A1:' + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    workbook.save(ReportExcelFile)

# EXCEL COLORS
for sheet in sheetsLits:
    worksheet = workbook[sheet]
    for cell in workbook[sheet][1]:
        worksheet[cell.coordinate].fill = PatternFill(fgColor = 'FFC6C1C1', fill_type = 'solid')
        workbook.save(ReportExcelFile)

# EXCEL COLUMN SIZE
for sheet in sheetsLits:
    for cell in workbook[sheet][1]:
        if get_column_letter(cell.column) == 'A':
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 30
        else:
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 10
        workbook.save(ReportExcelFile)


# CHART CATEGORIES
chartLabel = df_Categories_count['Categories'].tolist()
chartLegendLabel = df_Categories_count['Categories'].tolist()
chartValue = df_Categories_count['Total'].tolist()
chartLegendPercent = df_Categories_count['Percent'].tolist()

chartLabel[-1] = ''

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig2 = plt.figure()
plt.pie(chartValue, labels=chartLabel, colors=colors, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('Categories', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig2.savefig(workDirectory+'myplot2.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot2.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot2.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot2.png')
img.anchor = 'E4'

workbook['Categories'].add_image(img)
workbook.save(ReportExcelFile)


# CHART HOW DID YOU HEAR ABOUT AMS (CUSTOM FIELD How did you hear about AMS?)
chartLabel = df_HowDidYouHearAboutUs_count['How did you hear about AMS?'].tolist()
chartLegendLabel = df_HowDidYouHearAboutUs_count['How did you hear about AMS?'].tolist()
chartValue = df_HowDidYouHearAboutUs_count['Total'].tolist()
chartLegendPercent = df_HowDidYouHearAboutUs_count['Percent'].tolist()

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig3 = plt.figure(figsize=(14,10))
plt.pie(chartValue, labels=chartLabel, colors=colors, autopct='%1.1f%%', shadow=False, startangle=90)

plt.axis('equal')
plt.title('How did you hear about AMS (known)', pad=20, fontsize=18)

plt.legend(legendLabels, loc='best', fontsize=10)

fig3.savefig(workDirectory+'myplot3.png', dpi=70)
plt.clf()

im = Image.open(workDirectory+'myplot3.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot3.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot3.png')
img.anchor = 'E2'

workbook['How Did You Hear'].add_image(img)
workbook.save(ReportExcelFile)


# MAP COUNTRIES
df_Country_count.set_index('Live Location:Country', inplace=True)

my_values = df_Country_count['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_Country_count['Percent'] = np.digitize(my_values, my_range) - 1

map1 = plt.figure(figsize=(14, 8))

ax = map1.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_countries, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['COUNTRY_HB']
    if shp_ctry not in df_Country_count.index:
        color = '#dddddd'
    else:
        color = scheme[df_Country_count.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map1.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# Footer
plt.figtext(0.2, 0.17, WebinarFileName.replace('_', ' '), ha="left", fontsize=13, weight='bold')
plt.figtext(0.2, 0.14, 'Participants: '+str(participants)+' - New emails: '+str(newWebinar), ha="left", fontsize=11)

cb.remove()

map1.savefig(workDirectory+'mymap1.png', dpi=110, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap1.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap1.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap1.png')
img.anchor = 'E2'

workbook['Countries'].add_image(img)
workbook.save(ReportExcelFile)


# CHART NEW CATEGORIES
chartLabel = df_NewCategories_count['Categories'].tolist()
chartLegendLabel = df_NewCategories_count['Categories'].tolist()
chartValue = df_NewCategories_count['Total'].tolist()
chartLegendPercent = df_NewCategories_count['Percent'].tolist()

chartLabel[-1] = ''

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig4 = plt.figure()
plt.pie(chartValue, labels=chartLabel, colors=colors, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('Categories (new emails)', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig4.savefig(workDirectory+'myplot4.png', dpi=100)
plt.clf()

im = Image.open(workDirectory+'myplot4.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot4.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot4.png')
img.anchor = 'E4'

workbook['News Categories'].add_image(img)
workbook.save(ReportExcelFile)


# CHART NEW HOW DID YOU HEAR ABOUT AMS (CUSTOM FIELD How did you hear about AMS?)
chartLabel = df_NewHowDidYouHearAboutUs_count['How did you hear about AMS?'].tolist()
chartLegendLabel = df_NewHowDidYouHearAboutUs_count['How did you hear about AMS?'].tolist()
chartValue = df_NewHowDidYouHearAboutUs_count['Total'].tolist()
chartLegendPercent = df_NewHowDidYouHearAboutUs_count['Percent'].tolist()

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig5 = plt.figure(figsize=(14,10))
plt.pie(chartValue, labels=chartLabel, colors=colors, autopct='%1.1f%%', shadow=False, startangle=90)

plt.axis('equal')
plt.title('How did you hear about AMS (known, new emails)', pad=20, fontsize=18)

plt.legend(legendLabels, loc='best', fontsize=10)

fig5.savefig(workDirectory+'myplot5.png', dpi=70)
plt.clf()

im = Image.open(workDirectory+'myplot5.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot5.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot5.png')
img.anchor = 'E2'

workbook['News How Did You Hear'].add_image(img)
workbook.save(ReportExcelFile)


# MAP NEW COUNTRIES
df_NewCountry_count.set_index('Live Location:Country', inplace=True)

my_values = df_NewCountry_count['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_NewCountry_count['Percent'] = np.digitize(my_values, my_range) - 1

map2 = plt.figure(figsize=(14, 8))

ax = map2.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_countries, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['COUNTRY_HB']
    if shp_ctry not in df_NewCountry_count.index:
        color = '#dddddd'
    else:
        color = scheme[df_NewCountry_count.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map2.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# Footer
plt.figtext(0.2, 0.17, WebinarFileName.replace('_', ' ')+' (new emails)', ha="left", fontsize=13, weight='bold')

cb.remove()

map2.savefig(workDirectory+'mymap2.png', dpi=110, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap2.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap2.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap2.png')
img.anchor = 'E2'

workbook['News Countries'].add_image(img)
workbook.save(ReportExcelFile)


# CHART AREAS
chartLabel = df_AreasCount['continent_stat'].tolist()
chartLegendLabel = df_AreasCount['continent_stat'].tolist()
chartValue = df_AreasCount['Total'].tolist()
chartLegendPercent = df_AreasCount['Percent'].tolist()

chartLabel[-1] = ''

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig6 = plt.figure()
plt.pie(chartValue, labels=chartLabel, colors=colors, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('Areas', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig6.savefig(workDirectory+'myplot6.png', dpi=80)
plt.clf()

im = Image.open(workDirectory+'myplot6.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot6.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot6.png')
img.anchor = 'A13'

workbook['Areas'].add_image(img)
workbook.save(ReportExcelFile)


# MAP AREAS
df_AreasCount.set_index('continent_stat', inplace=True)

my_values = df_AreasCount['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_AreasCount['Percent'] = np.digitize(my_values, my_range) - 1

map3 = plt.figure(figsize=(14, 8))

ax = map3.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_areas, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['continent']
    if shp_ctry not in df_AreasCount.index:
        color = '#dddddd'
    else:
        color = scheme[df_AreasCount.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map3.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# Footer
plt.figtext(0.2, 0.17, WebinarFileName.replace('_', ' '), ha="left", fontsize=13, weight='bold')
plt.figtext(0.2, 0.14, 'Participants: '+str(participants)+' - New emails: '+str(newWebinar), ha="left", fontsize=11)

cb.remove()

map3.savefig(workDirectory+'mymap3.png', dpi=90, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap3.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap3.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap3.png')
img.anchor = 'G2'

workbook['Areas'].add_image(img)
workbook.save(ReportExcelFile)


# CHART NEW AREAS
chartLabel = df_NewAreasCount['continent_stat'].tolist()
chartLegendLabel = df_NewAreasCount['continent_stat'].tolist()
chartValue = df_NewAreasCount['Total'].tolist()
chartLegendPercent = df_NewAreasCount['Percent'].tolist()

chartLabel[-1] = ''

legendLabels = []
for i, j in zip(chartLegendLabel, map(str, chartLegendPercent)):
    legendLabels.append(i + ' (' + j + ' %)')

colors = plt.rcParams['axes.prop_cycle'].by_key()['color']

fig7 = plt.figure()
plt.pie(chartValue, labels=chartLabel, colors=colors, autopct=None, shadow=False, startangle=90)
plt.axis('equal')
plt.title('Areas', pad=20, fontsize=15)

plt.legend(legendLabels, loc='best', fontsize=8)

fig7.savefig(workDirectory+'myplot7.png', dpi=80)
plt.clf()

im = Image.open(workDirectory+'myplot7.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'myplot7.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'myplot7.png')
img.anchor = 'A13'

workbook['News Areas'].add_image(img)
workbook.save(ReportExcelFile)


# MAP NEW AREAS
df_NewAreasCount.set_index('continent_stat', inplace=True)

my_values = df_NewAreasCount['Percent']

num_colors = 30
cm = plt.get_cmap('Blues')
scheme = [cm(i / num_colors) for i in range(num_colors)]

my_range = np.linspace(my_values.min(), my_values.max(), num_colors)

df_NewAreasCount['Percent'] = np.digitize(my_values, my_range) - 1

map4 = plt.figure(figsize=(14, 8))

ax = map4.add_subplot(111, frame_on=False)

m = Basemap(lon_0=0, projection='robin')
m.drawmapboundary(color='w')

m.readshapefile(shp_simple_areas, 'units', color='#444444', linewidth=.2, default_encoding='iso-8859-15')

for info, shape in zip(m.units_info, m.units):
    shp_ctry = info['continent']
    if shp_ctry not in df_NewAreasCount.index:
        color = '#dddddd'
    else:
        color = scheme[df_NewAreasCount.loc[shp_ctry]['Percent']]

    patches = [Polygon(np.array(shape), True)]
    pc = PatchCollection(patches)
    pc.set_facecolor(color)
    ax.add_collection(pc)

# Cover up Antarctica
ax.axhspan(0, 1000 * 1800, facecolor='w', edgecolor='w', zorder=2)

# Draw color legend
ax_legend = map4.add_axes([0.2, 0.14, 0.6, 0.03], zorder=3)
cmap = mpl.colors.ListedColormap(scheme)
cb = mpl.colorbar.ColorbarBase(ax_legend, cmap=cmap, ticks=my_range, boundaries=my_range, orientation='horizontal')

# Footer
plt.figtext(0.2, 0.17, WebinarFileName.replace('_', ' ')+' (new emails)', ha="left", fontsize=13, weight='bold')

cb.remove()

map4.savefig(workDirectory+'mymap4.png', dpi=90, bbox_inches='tight')
plt.clf()

im = Image.open(workDirectory+'mymap4.png')
bordered = ImageOps.expand(im, border=1, fill=(0, 0, 0))
bordered.save(workDirectory+'mymap4.png')

# INSERT IN EXCEL
img = openpyxl.drawing.image.Image(workDirectory+'mymap4.png')
img.anchor = 'G2'

workbook['News Areas'].add_image(img)
workbook.save(ReportExcelFile)


# REMOVE PICTURES
os.remove(workDirectory+'myplot2.png')
os.remove(workDirectory+'myplot3.png')
os.remove(workDirectory+'mymap1.png')
os.remove(workDirectory+'myplot4.png')
os.remove(workDirectory+'myplot5.png')
os.remove(workDirectory+'mymap2.png')
os.remove(workDirectory+'mymap3.png')
os.remove(workDirectory+'myplot6.png')
os.remove(workDirectory+'mymap4.png')
os.remove(workDirectory+'myplot7.png')


# TERMINAL OUTPUTS AND TESTS
print("OK, export done!")